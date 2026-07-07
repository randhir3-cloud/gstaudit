import os
import re
import copy as _copy_module
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
import glob
import openpyxl
from openpyxl.styles import PatternFill, Font as OxlFont

# ---------------------------------------------------------------------------
# Month lookup for GSTR-1 period extraction from filename
# ---------------------------------------------------------------------------
MONTH_MAP = {
    '01': 'Jan', '02': 'Feb', '03': 'Mar', '04': 'Apr',
    '05': 'May', '06': 'Jun', '07': 'Jul', '08': 'Aug',
    '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dec',
}

# Full month names — used when building the Tax Period range in Read me
FULL_MONTH_MAP = {
    '01': 'January',  '02': 'February', '03': 'March',    '04': 'April',
    '05': 'May',      '06': 'June',     '07': 'July',     '08': 'August',
    '09': 'September','10': 'October',  '11': 'November', '12': 'December',
}

GSTR1_SKIP_SHEETS = {'read me'}          # sheet names to exclude from data merge (lower-case)
GSTR1_HEADER_ROW  = 3                    # 0-indexed → row 4 in Excel is the actual header

# Read me row indices (0-based) and the column that holds the value
README_TAX_PERIOD_ROW = 4   # Row 5  → "Tax Period"
README_ARN_ROW        = 8   # Row 9  → "ARN"
README_ARN_DATE_ROW   = 9   # Row 10 → "ARN date"
README_VALUE_COL      = 2   # Column C (0-based index 2) holds the value for each label


# ===========================================================================
# LAUNCHER — Home screen
# ===========================================================================
class LauncherApp:
    """Main menu that lets the user pick which merger to open."""

    def __init__(self, master):
        self.master = master
        master.title("Excel Merger App ::: by Randhir Singh")
        master.geometry("640x360")
        master.minsize(600, 320)
        master.resizable(True, True)

        # ---- Header --------------------------------------------------------
        header = tk.Frame(master, bg="#1e3a5f", height=70)
        header.pack(fill='x')
        header.pack_propagate(False)

        tk.Label(
            header,
            text="Excel Merger Tool",
            font=("Segoe UI", 18, "bold"),
            fg="white", bg="#1e3a5f"
        ).pack(expand=True)

        # ---- Subtitle ------------------------------------------------------
        tk.Label(
            master,
            text="Select the merge mode to continue:",
            font=("Segoe UI", 11),
            fg="#444444"
        ).pack(pady=(24, 12))

        # ---- Buttons (stretch with window) ---------------------------------
        btn_frame = tk.Frame(master)
        btn_frame.pack(fill='x', expand=True, padx=40, pady=10)
        btn_frame.columnconfigure(0, weight=1)
        btn_frame.columnconfigure(1, weight=1)

        eway_btn = tk.Button(
            btn_frame,
            text="📋  E-Way Bill Merge",
            font=("Segoe UI", 12, "bold"),
            bg="#2e7d32", fg="white",
            activebackground="#1b5e20", activeforeground="white",
            relief="flat", padx=20, pady=16,
            cursor="hand2",
            command=self.open_eway
        )
        eway_btn.grid(row=0, column=0, padx=12, sticky='ew')

        gstr_btn = tk.Button(
            btn_frame,
            text="🧾  GSTR-1 Merge",
            font=("Segoe UI", 12, "bold"),
            bg="#1565c0", fg="white",
            activebackground="#0d47a1", activeforeground="white",
            relief="flat", padx=20, pady=16,
            cursor="hand2",
            command=self.open_gstr1
        )
        gstr_btn.grid(row=0, column=1, padx=12, sticky='ew')

        # ---- Footer --------------------------------------------------------
        tk.Label(
            master,
            text="by Randhir Singh",
            font=("Segoe UI", 9),
            fg="#aaaaaa"
        ).pack(side='bottom', pady=8)

    def open_eway(self):
        self.master.withdraw()
        win = tk.Toplevel(self.master)
        win.protocol("WM_DELETE_WINDOW", lambda: self._on_child_close(win))
        EWayBillMergerApp(win)

    def open_gstr1(self):
        self.master.withdraw()
        win = tk.Toplevel(self.master)
        win.protocol("WM_DELETE_WINDOW", lambda: self._on_child_close(win))
        GSTR1MergerApp(win)

    def _on_child_close(self, win):
        """Return to launcher when a child window is closed."""
        win.destroy()
        self.master.deiconify()


# ===========================================================================
# E-WAY BILL MERGER  (original logic — untouched)
# ===========================================================================
class EWayBillMergerApp:
    """Merges all sheets from all selected files into one flat sheet."""

    def __init__(self, master):
        self.master = master
        master.title("E-Way Bill Merge ::: by Randhir Singh")
        master.geometry("800x600")

        # ---- Top bar -------------------------------------------------------
        top_frame = ttk.Frame(master, padding="10")
        top_frame.pack(fill='x')

        folder_btn = ttk.Button(top_frame, text="Select Folder", command=self.select_folder)
        folder_btn.pack(side='left')

        self.output_filename = tk.StringVar(value="eway_merged_output.xlsx")
        ttk.Label(top_frame, text="Output Filename:").pack(side='left', padx=(10, 2))
        ttk.Entry(top_frame, textvariable=self.output_filename, width=30).pack(side='left')

        # ---- File list area ------------------------------------------------
        self.files_frame = ttk.Frame(master, padding="10")
        self.files_frame.pack(fill='both', expand=True)

        self.canvas = tk.Canvas(self.files_frame)
        self.scrollbar = ttk.Scrollbar(self.files_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        ttk.Button(master, text="Merge Files", command=self.merge_files).pack(pady=10)

        self.file_entries = []

    # ---- Folder selection --------------------------------------------------
    def select_folder(self):
        folder_selected = filedialog.askdirectory()
        if not folder_selected:
            return

        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.file_entries = []

        files_xlsx = glob.glob(os.path.join(folder_selected, "*.xlsx"))
        files_xls  = glob.glob(os.path.join(folder_selected, "*.xls"))

        converted_files = []
        for file in files_xls:
            new_file = file[:-4] + '.xlsx'
            try:
                with open(file, 'rb') as f:
                    header = f.read(500).lower()
                if b'<html' in header or b'<!doctype html' in header or b'<style' in header:
                    tables = pd.read_html(file)
                    if not tables:
                        raise ValueError("No tables found in HTML file")
                    with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
                        tables[0].to_excel(writer, sheet_name="Sheet1", index=False)
                else:
                    xls = pd.ExcelFile(file, engine='xlrd')
                    with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
                        for sheet_name in xls.sheet_names:
                            df = pd.read_excel(file, sheet_name=sheet_name, engine='xlrd')
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                converted_files.append(new_file)
            except Exception as e:
                messagebox.showerror("Conversion Error",
                    f"Error converting {os.path.basename(file)} to .xlsx:\n{str(e)}")

        files = sorted(files_xlsx + converted_files)
        if not files:
            messagebox.showinfo("No Files", "No Excel files found in the selected folder.")
            return

        for file in files:
            self.file_entries.append({'file': file, 'var': tk.BooleanVar(value=True)})
        self.render_file_list()

    # ---- UI helpers --------------------------------------------------------
    def render_file_list(self):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        for index, entry in enumerate(self.file_entries):
            frame = ttk.Frame(self.scrollable_frame)
            frame.pack(fill='x', pady=2, padx=2)
            ttk.Checkbutton(frame, text=os.path.basename(entry['file']),
                            variable=entry['var']).pack(side="left", fill='x', expand=True)
            btn_up = ttk.Button(frame, text="Up", command=lambda idx=index: self.move_up(idx))
            btn_up.pack(side="left", padx=2)
            if index == 0:
                btn_up.state(["disabled"])
            btn_down = ttk.Button(frame, text="Down", command=lambda idx=index: self.move_down(idx))
            btn_down.pack(side="left", padx=2)
            if index == len(self.file_entries) - 1:
                btn_down.state(["disabled"])

    def move_up(self, index):
        if index > 0:
            self.file_entries[index], self.file_entries[index - 1] = \
                self.file_entries[index - 1], self.file_entries[index]
            self.render_file_list()

    def move_down(self, index):
        if index < len(self.file_entries) - 1:
            self.file_entries[index], self.file_entries[index + 1] = \
                self.file_entries[index + 1], self.file_entries[index]
            self.render_file_list()

    # ---- Merge logic -------------------------------------------------------
    def merge_files(self):
        if not self.file_entries:
            messagebox.showerror("Error", "No files selected! Please select a folder first.")
            return

        selected_files = [e['file'] for e in self.file_entries if e['var'].get()]
        if not selected_files:
            messagebox.showerror("Error", "No files selected for merging!")
            return

        output_file = self.output_filename.get()
        if not output_file.endswith('.xlsx'):
            output_file += '.xlsx'

        try:
            all_dfs = []
            for file in selected_files:
                try:
                    xls = pd.ExcelFile(file, engine='openpyxl')
                    for sheet_name in xls.sheet_names:
                        try:
                            df = pd.read_excel(file, sheet_name=sheet_name, engine='openpyxl')
                        except Exception:
                            df = pd.read_excel(file, sheet_name=sheet_name)
                        df['Source_File']  = os.path.basename(file)
                        df['Source_Sheet'] = sheet_name
                        all_dfs.append(df)
                except Exception as e:
                    messagebox.showerror("File Error",
                        f"Error processing {os.path.basename(file)}:\n{str(e)}")
                    return

            if not all_dfs:
                messagebox.showerror("Error", "No data found to merge!")
                return

            combined_df = pd.concat(all_dfs, ignore_index=True)
            output_path = os.path.join(os.path.dirname(selected_files[0]), output_file)
            combined_df.to_excel(output_path, index=False, engine='openpyxl')

            messagebox.showinfo("Success",
                f"Merged {len(selected_files)} files successfully!\nOutput saved to:\n{output_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")


# ===========================================================================
# GSTR-1 MERGER  (new — merges same-named sheets across monthly files)
# ===========================================================================
class GSTR1MergerApp:
    """
    Merges GSTR-1 monthly Excel files.

    Each source file has identical sheets (20 sheets). For every sheet
    except 'Read me', data rows from all selected files are stacked
    into one sheet in the output workbook.

    Header is at row 4 in each sheet (rows 1-3 are GSTR title text).
    A 'Source_Period' column (e.g. 'Jan-2023') is added to each row,
    extracted from the filename pattern  ..._MMYYYY_...
    """

    def __init__(self, master):
        self.master = master
        master.title("GSTR-1 Merge ::: by Randhir Singh")
        master.geometry("820x640")

        # ---- Header band ---------------------------------------------------
        header = tk.Frame(master, bg="#1565c0", height=50)
        header.pack(fill='x')
        tk.Label(
            header,
            text="🧾  GSTR-1 Monthly Merger",
            font=("Segoe UI", 14, "bold"),
            fg="white", bg="#1565c0"
        ).pack(side='left', padx=14, pady=10)

        # ---- Top controls --------------------------------------------------
        top_frame = ttk.Frame(master, padding="10")
        top_frame.pack(fill='x')

        ttk.Button(top_frame, text="📂  Select Folder", command=self.select_folder).pack(side='left')

        self.output_filename = tk.StringVar(value="GSTR1_MERGED.xlsx")
        ttk.Label(top_frame, text="  Output File:").pack(side='left')
        ttk.Entry(top_frame, textvariable=self.output_filename, width=36).pack(side='left', padx=4)

        # ---- Info label ----------------------------------------------------
        self.info_var = tk.StringVar(value="No folder selected")
        ttk.Label(master, textvariable=self.info_var,
                  font=("Segoe UI", 9), foreground="#555555").pack(anchor='w', padx=14)

        # ---- File list with scrollbar --------------------------------------
        self.files_frame = ttk.LabelFrame(master, text="Files to merge", padding="6")
        self.files_frame.pack(fill='both', expand=True, padx=10, pady=6)

        self.canvas = tk.Canvas(self.files_frame)
        self.scrollbar = ttk.Scrollbar(self.files_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # ---- Sheets-to-skip note -------------------------------------------
        note = ttk.Label(master,
            text="ℹ  'Read me' sheet will be automatically skipped.",
            font=("Segoe UI", 9, "italic"), foreground="#777777")
        note.pack(anchor='w', padx=14)

        # ---- Merge button --------------------------------------------------
        tk.Button(
            master,
            text="🔀  Merge GSTR-1 Files",
            font=("Segoe UI", 12, "bold"),
            bg="#1565c0", fg="white",
            activebackground="#0d47a1", activeforeground="white",
            relief="flat", padx=16, pady=10,
            cursor="hand2",
            command=self.merge_files
        ).pack(pady=10)

        self.file_entries = []
        self.folder_path  = ""

    # ---- Period extraction -------------------------------------------------
    @staticmethod
    def extract_period(filename: str) -> str:
        """
        Extract a human-readable period from the filename.
        Expects a segment like '012023' or '042022' (MMYYYY) anywhere in the name.
        Returns e.g. 'Jan-2023' or the raw segment if pattern not found.
        """
        basename = os.path.basename(filename)
        match = re.search(r'_(\d{2})(\d{4})_', basename)
        if match:
            mm, yyyy = match.group(1), match.group(2)
            return f"{MONTH_MAP.get(mm, mm)}-{yyyy}"
        # fallback: return full basename without extension
        return os.path.splitext(basename)[0]

    # ---- Financial-year helpers --------------------------------------------
    @staticmethod
    def _fy_sort_key(mm: str, yyyy: str) -> int:
        """
        Convert a calendar month/year pair to an integer sort key that
        orders months in Indian Financial Year sequence (April = 1 … March = 12).

        The key is encoded as  FY_START_YEAR * 100 + FY_MONTH_INDEX, e.g.:
          Apr-2022 → 202201,  Dec-2022 → 202209,
          Jan-2023 → 202210,  Mar-2023 → 202212
        """
        mm_int, yyyy_int = int(mm), int(yyyy)
        if mm_int >= 4:
            return yyyy_int * 100 + (mm_int - 3)   # Apr→1 … Dec→9
        else:
            return (yyyy_int - 1) * 100 + (mm_int + 9)  # Jan→10 … Mar→12

    @staticmethod
    def _fy_key_to_label(key: int) -> str:
        """Convert an _fy_sort_key integer back to a human label like 'January 2023'."""
        idx          = key % 100          # 1-12
        fy_start_yr  = key // 100
        if idx <= 9:                       # April … December
            cal_month = idx + 3
            cal_year  = fy_start_yr
        else:                              # January … March
            cal_month = idx - 9
            cal_year  = fy_start_yr + 1
        return f"{FULL_MONTH_MAP.get(f'{cal_month:02d}', str(cal_month))} {cal_year}"

    @staticmethod
    def _next_fy_key(key: int) -> int:
        """Advance an _fy_sort_key by exactly one month."""
        idx     = key % 100
        fy_year = key // 100
        if idx < 12:
            return fy_year * 100 + (idx + 1)
        return (fy_year + 1) * 100 + 1

    def _file_fy_key(self, filepath: str) -> int:
        """Return the FY sort key for a file path (999999 if pattern not found)."""
        m = re.search(r'_(\d{2})(\d{4})_', os.path.basename(filepath))
        if m:
            return self._fy_sort_key(m.group(1), m.group(2))
        return 999999

    def _find_missing_months(self, entries: list) -> list:
        """
        Given a list of file-entry dicts, return a list of human-readable
        month labels (e.g. ['June 2022', 'October 2022']) that are missing
        between the first and last file's month in FY order.
        Returns an empty list when no gap is detected or < 2 files matched.
        """
        present = set()
        for entry in entries:
            m = re.search(r'_(\d{2})(\d{4})_', os.path.basename(entry['file']))
            if m:
                present.add(self._fy_sort_key(m.group(1), m.group(2)))

        if len(present) < 2:
            return []

        min_key = min(present)
        max_key = max(present)
        missing = []
        cur = self._next_fy_key(min_key)
        while cur < max_key:
            if cur not in present:
                missing.append(self._fy_key_to_label(cur))
            cur = self._next_fy_key(cur)
        return missing

    # ---- Folder selection --------------------------------------------------
    def select_folder(self):
        folder_selected = filedialog.askdirectory()
        if not folder_selected:
            return

        self.folder_path = folder_selected

        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.file_entries = []

        files = sorted(glob.glob(os.path.join(folder_selected, "*.xlsx")) +
                       glob.glob(os.path.join(folder_selected, "*.xls")))

        # Exclude any previously merged output from the list
        files = [f for f in files if "MERGED" not in os.path.basename(f).upper()
                 and "_Merge" not in os.path.basename(f)]

        if not files:
            messagebox.showinfo("No Files", "No Excel files found in the selected folder.")
            self.info_var.set("No files found.")
            return

        for file in files:
            period = self.extract_period(file)
            self.file_entries.append({
                'file':   file,
                'period': period,
                'var':    tk.BooleanVar(value=True)
            })

        # Sort entries in Financial Year order (April → March)
        self.file_entries.sort(key=lambda e: self._file_fy_key(e['file']))

        self.info_var.set(f"Folder: {folder_selected}   |   {len(files)} file(s) found")
        self.render_file_list()

    # ---- UI helpers --------------------------------------------------------
    def render_file_list(self):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        for index, entry in enumerate(self.file_entries):
            frame = ttk.Frame(self.scrollable_frame)
            frame.pack(fill='x', pady=2, padx=2)

            label = f"{os.path.basename(entry['file'])}   [{entry['period']}]"
            ttk.Checkbutton(frame, text=label, variable=entry['var']).pack(
                side="left", fill='x', expand=True)

            btn_up = ttk.Button(frame, text="▲", width=3,
                                command=lambda idx=index: self.move_up(idx))
            btn_up.pack(side="left", padx=2)
            if index == 0:
                btn_up.state(["disabled"])

            btn_down = ttk.Button(frame, text="▼", width=3,
                                  command=lambda idx=index: self.move_down(idx))
            btn_down.pack(side="left", padx=2)
            if index == len(self.file_entries) - 1:
                btn_down.state(["disabled"])

    def move_up(self, index):
        if index > 0:
            self.file_entries[index], self.file_entries[index - 1] = \
                self.file_entries[index - 1], self.file_entries[index]
            self.render_file_list()

    def move_down(self, index):
        if index < len(self.file_entries) - 1:
            self.file_entries[index], self.file_entries[index + 1] = \
                self.file_entries[index + 1], self.file_entries[index]
            self.render_file_list()

    # ---- Style helpers -----------------------------------------------------
    @staticmethod
    def _copy_cell_style(src, dst):
        """
        Copy ALL formatting attributes (font, fill, border, alignment,
        number_format, protection) from openpyxl cell *src* to *dst*.
        Used for HEADER-ROW cells only — keeps the coloured header intact.
        Safe to call when src is None or has no style.
        """
        if src is None or not src.has_style:
            return
        dst.font          = _copy_module.copy(src.font)
        dst.fill          = _copy_module.copy(src.fill)
        dst.border        = _copy_module.copy(src.border)
        dst.alignment     = _copy_module.copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection    = _copy_module.copy(src.protection)

    @staticmethod
    def _apply_data_style(src_hdr, dst):
        """
        Apply style for a DATA row cell (rows 5+).
        Only inherits number_format from the corresponding header cell so
        that dates/currency/percentages display correctly.
        Fill is explicitly reset to no-fill (white/default background) and
        font colour is reset to solid black so data rows never inherit the
        dark-blue header styling from the template workbook.
        """
        if src_hdr is not None and src_hdr.has_style:
            dst.number_format = src_hdr.number_format
        # Explicitly reset to default white background
        dst.fill = PatternFill(fill_type=None)
        # Explicitly reset to default black text
        dst.font = OxlFont(color='FF000000')

    @staticmethod
    def _safe_value(val):
        """
        Convert pandas/numpy special values to Python-native types that
        openpyxl can write safely (NaN/NaT → None, numpy scalars → Python).
        """
        try:
            if pd.isna(val):
                return None
        except (TypeError, ValueError):
            pass
        if isinstance(val, np.integer):
            return int(val)
        if isinstance(val, np.floating):
            return float(val)
        if isinstance(val, np.bool_):
            return bool(val)
        return val

    # ---- Merge logic -------------------------------------------------------
    def merge_files(self):
        if not self.file_entries:
            messagebox.showerror("Error", "No files selected! Please select a folder first.")
            return

        selected = [e for e in self.file_entries if e['var'].get()]
        if not selected:
            messagebox.showerror("Error", "No files checked for merging!")
            return

        # ---- Missing-month check -------------------------------------------
        missing = self._find_missing_months(selected)
        if missing:
            month_list = "\n".join(f"  • {m}" for m in missing)
            answer = messagebox.askyesno(
                "⚠️  Missing Months Detected",
                f"The following month(s) are missing between your selected files:\n\n"
                f"{month_list}\n\n"
                f"Do you want to continue merging without the missing file(s)?\n"
                f"  Yes → continue merge\n"
                f"  No  → cancel so you can add the missing file(s)",
                icon='warning'
            )
            if not answer:
                return   # user chose to cancel and add missing files

        try:
            # ----------------------------------------------------------------
            # Step 1: Determine Tax Period range from selected filenames
            #         e.g. "April 2022 to March 2023"
            # ----------------------------------------------------------------
            period_keys = []   # (sort_int, full_month_name, year_str)
            for entry in selected:
                m = re.search(r'_(\d{2})(\d{4})_', os.path.basename(entry['file']))
                if m:
                    mm, yyyy = m.group(1), m.group(2)
                    period_keys.append(
                        (int(yyyy) * 100 + int(mm), FULL_MONTH_MAP.get(mm, mm), yyyy)
                    )

            if period_keys:
                period_keys.sort()
                first_label = f"{period_keys[0][1]} {period_keys[0][2]}"
                last_label  = f"{period_keys[-1][1]} {period_keys[-1][2]}"
                tax_period_text = (
                    first_label if first_label == last_label
                    else f"{first_label} to {last_label}"
                )
            else:
                tax_period_text = "Full Period"

            # ----------------------------------------------------------------
            # Step 2: Collect data rows per sheet from ALL selected files.
            #         Sheets with zero data rows across all files are simply
            #         left blank — their template formatting is kept intact.
            # ----------------------------------------------------------------
            sheet_data: dict[str, list] = {}   # sheet_name → [DataFrame, …]

            for entry in selected:
                file   = entry['file']
                period = entry['period']
                try:
                    xls = pd.ExcelFile(file, engine='openpyxl')
                except Exception:
                    xls = pd.ExcelFile(file)

                for sheet_name in xls.sheet_names:
                    if sheet_name.strip().lower() in GSTR1_SKIP_SHEETS:
                        continue

                    try:
                        df = pd.read_excel(
                            file, sheet_name=sheet_name,
                            header=GSTR1_HEADER_ROW, engine='openpyxl'
                        )
                    except Exception:
                        df = pd.read_excel(
                            file, sheet_name=sheet_name,
                            header=GSTR1_HEADER_ROW
                        )

                    df.dropna(how='all', inplace=True)
                    if df.empty:
                        continue   # no data rows this month — sheet stays blank

                    df.insert(0, 'Source_Period', period)
                    if sheet_name not in sheet_data:
                        sheet_data[sheet_name] = []
                    sheet_data[sheet_name].append(df)

            # ----------------------------------------------------------------
            # Step 3: Load the first selected file as the formatting TEMPLATE.
            #         This gives us all 20 sheets with merged cells, column
            #         widths, row heights, freeze panes, tab colours, etc.
            # ----------------------------------------------------------------
            first_file = selected[0]['file']
            wb = openpyxl.load_workbook(first_file)

            # 1-based Excel row indices
            HEADER_ROW = GSTR1_HEADER_ROW + 1   # row 4  (header)
            DATA_START = HEADER_ROW + 1           # row 5  (first data row)

            # ----------------------------------------------------------------
            # Step 3b: Auto-generate output filename from Read me metadata.
            #          Format: GSTR1_{GSTIN}_{FinancialYear}_Merge.xlsx
            # ----------------------------------------------------------------
            readme_ws_name = next(
                (sn for sn in wb.sheetnames if sn.strip().lower() == 'read me'),
                None
            )
            val_col = README_VALUE_COL + 1   # 0-based → 1-based (col C = 3)

            if readme_ws_name:
                ws_rm_meta = wb[readme_ws_name]
                gstin_val  = ws_rm_meta.cell(row=6, column=val_col).value or ''
                fy_val     = ws_rm_meta.cell(row=4, column=val_col).value or ''
                safe_gstin = re.sub(r'[\\/:*?"<>|]', '_', str(gstin_val).strip())
                safe_fy    = re.sub(r'[\\/:*?"<>|]', '_', str(fy_val).strip())
                auto_name  = f"GSTR1_{safe_gstin}_{safe_fy}_Merged.xlsx"
            else:
                auto_name = "GSTR1_Merged.xlsx"

            output_path = os.path.join(
                self.folder_path or os.path.dirname(selected[0]['file']),
                auto_name
            )
            self.output_filename.set(auto_name)   # reflect in the UI text box

            # ----------------------------------------------------------------
            # Step 4: Patch "Read me" Tax Period / ARN directly in the
            #         template worksheet — preserves all merged cells and
            #         complex formatting of that sheet.
            # ----------------------------------------------------------------
            if readme_ws_name:
                ws_rm = wb[readme_ws_name]
                ws_rm.cell(
                    row=README_TAX_PERIOD_ROW + 1, column=val_col
                ).value = tax_period_text
                ws_rm.cell(row=README_ARN_ROW      + 1, column=val_col).value = ''
                ws_rm.cell(row=README_ARN_DATE_ROW + 1, column=val_col).value = ''

            # ----------------------------------------------------------------
            # Step 5: For every sheet in the template workbook:
            #   • If the sheet has no collected data  → leave it untouched
            #     (blank sheet preserved with full original formatting).
            #   • If the sheet has data → clear rows 5+ then rewrite the
            #     merged, period-tagged rows.
            #
            #   Rows 1-4 of each sheet are NEVER touched — their dark-blue
            #   header formatting is inherited intact from the template.
            #
            #   Data rows (row 5+) use DEFAULT cell style (white background,
            #   black text), inheriting only number_format from the header
            #   so dates/currency display correctly.
            #
            #   Source_Period is written as the LAST column (after all
            #   original columns) so the original column layout in rows 1-4
            #   is never disturbed.
            # ----------------------------------------------------------------
            total_sheets     = len(wb.sheetnames)
            sheets_with_data = 0

            for sheet_name in wb.sheetnames:
                if sheet_name.strip().lower() in GSTR1_SKIP_SHEETS:
                    continue

                # No data collected → keep template sheet exactly as-is
                if sheet_name not in sheet_data:
                    continue

                sheets_with_data += 1
                ws = wb[sheet_name]

                # Build header-name → column-number map from template row 4
                template_col_map: dict[str, int] = {}
                max_data_col = 0
                for cell in ws[HEADER_ROW]:
                    if cell.value is not None:
                        template_col_map[str(cell.value).strip()] = cell.column
                        max_data_col = max(max_data_col, cell.column)

                # Cache header-row cell objects (for number_format reference)
                hdr_cells: dict[int, object] = {
                    cell.column: cell for cell in ws[HEADER_ROW]
                }

                # --- Clear existing data rows (row 5 to max_row) ---
                if ws.max_row >= DATA_START:
                    ws.delete_rows(DATA_START, ws.max_row - DATA_START + 1)

                # --- Concatenate all periods' DataFrames for this sheet ---
                combined = pd.concat(sheet_data[sheet_name], ignore_index=True)

                # --- Source_Period column: one after the last original col ---
                sp_col     = max_data_col + 1
                sp_ref_hdr = hdr_cells.get(max_data_col)   # for style reference

                # Add "Source_Period" header in row 4 (coloured like other headers)
                sp_hdr_cell = ws.cell(
                    row=HEADER_ROW, column=sp_col, value='Source_Period'
                )
                self._copy_cell_style(sp_ref_hdr, sp_hdr_cell)

                # --- Write data rows (white background, black text) ---
                for r_offset, (_, row) in enumerate(combined.iterrows()):
                    excel_row = DATA_START + r_offset

                    # Write each original column matched by header name
                    for col_name, col_num in template_col_map.items():
                        raw = row.get(col_name)
                        val = self._safe_value(raw)
                        cell = ws.cell(row=excel_row, column=col_num, value=val)
                        # Only copy number_format — fill/font stay as default
                        self._apply_data_style(hdr_cells.get(col_num), cell)

                    # Write Source_Period value in the extra column
                    sp_cell = ws.cell(
                        row=excel_row, column=sp_col,
                        value=row.get('Source_Period', '')
                    )
                    self._apply_data_style(sp_ref_hdr, sp_cell)

            # ----------------------------------------------------------------
            # Step 6: Save the modified template workbook as the output file
            # ----------------------------------------------------------------
            wb.save(output_path)

            blank_count = total_sheets - 1 - sheets_with_data  # excl. Read me
            messagebox.showinfo(
                "✅  Merge Complete",
                f"Merged {len(selected)} file(s) successfully.\n"
                f"All {total_sheets} sheet(s) written "
                f"({sheets_with_data} with data, {blank_count} blank).\n"
                f"Tax Period set to: {tax_period_text}\n\n"
                f"Output saved to:\n{output_path}"
            )

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during merge:\n{str(e)}")


# ===========================================================================
# Entry point
# ===========================================================================
if __name__ == "__main__":
    root = tk.Tk()
    app  = LauncherApp(root)
    root.mainloop()