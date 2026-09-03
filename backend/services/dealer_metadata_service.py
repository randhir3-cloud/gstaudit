"""Extract dealer metadata from GSTR Read me worksheets."""

from __future__ import annotations

import io
import re
from typing import Dict, List, Optional, Tuple, Any

import openpyxl

from models.dealer_metadata import DealerMetadata

README_SHEET_NAMES = {"read me", "readme", "read_me"}

FIELD_ALIASES: Dict[str, Tuple[str, ...]] = {
    "gstin": ("gstin", "taxpayer's gstin", "taxpayers gstin", "taxpayer gstin"),
    "legal_name": ("legal name", "legal name of taxpayer", "legalname"),
    "trade_name": ("trade name", "trade name (if any)", "tradename"),
    "financial_year": ("financial year", "financialyear", "fy"),
    "tax_period": ("tax period", "tax period ", "taxperiod"),
    "arn": ("arn",),
    "arn_date": ("arn date",),
    "download_date": (
        "date and time of generation",
        "date of generation",
        "download date",
        "date of download",
    ),
}

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

MONTH_NAME_MAP = {
    "january": 1, "jan": 1, "01": 1, "1": 1,
    "february": 2, "feb": 2, "02": 2, "2": 2,
    "march": 3, "mar": 3, "03": 3, "3": 3,
    "april": 4, "apr": 4, "04": 4, "4": 4,
    "may": 5, "05": 5, "5": 5,
    "june": 6, "jun": 6, "06": 6, "6": 6,
    "july": 7, "jul": 7, "07": 7, "7": 7,
    "august": 8, "aug": 8, "08": 8, "8": 8,
    "september": 9, "sep": 9, "sept": 9, "09": 9, "9": 9,
    "october": 10, "oct": 10, "10": 10,
    "november": 11, "nov": 11, "11": 11,
    "december": 12, "dec": 12, "12": 12,
}

# Known GST portal cell positions (1-based row, col) when label scan fails.
GSTR1_FALLBACK: Dict[str, Tuple[int, int]] = {
    "financial_year": (4, 3),
    "tax_period": (5, 3),
    "gstin": (6, 3),
    "legal_name": (7, 3),
    "trade_name": (8, 3),
    "arn": (9, 3),
    "arn_date": (10, 3),
    "download_date": (11, 3),
}

GSTR2A_FALLBACK: Dict[str, Tuple[int, int]] = {
    "gstin": (2, 3),
    "legal_name": (3, 3),
    "trade_name": (4, 3),
    "tax_period": (2, 5),
    "financial_year": (3, 5),
    "download_date": (4, 5),
    "arn": (0, 0),
    "arn_date": (0, 0),
}


def _normalize_label(text: object) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text).strip().lower())


def _clean_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def find_readme_sheet_name(sheet_names: List[str]) -> Optional[str]:
    for name in sheet_names:
        if name.strip().lower() in README_SHEET_NAMES:
            return name
    return None


def _match_field(label: str) -> Optional[str]:
    normalized = _normalize_label(label)
    if not normalized:
        return None
    for field, aliases in FIELD_ALIASES.items():
        if normalized in aliases:
            return field
        for alias in aliases:
            if alias in normalized or normalized in alias:
                return field
    return None


def parse_financial_year(raw_fy: object) -> Tuple[Optional[int], Optional[int], str]:
    text = _clean_value(raw_fy)
    if not text:
        return None, None, ""
    m = re.match(r"^(\d{4})[-/](\d{2,4})$", text)
    if m:
        start_year = int(m.group(1))
        end_year = int(m.group(2))
        if end_year < 100:
            end_year = (start_year // 100) * 100 + end_year
        short_end = str(end_year)[-2:]
        return start_year, end_year, f"{start_year}-{short_end}"
    return None, None, text


def parse_gstr2a_tax_period(raw_period: object, financial_year: str = "") -> Tuple[Optional[int], Optional[int], str, str]:
    if raw_period is None:
        return None, None, "", ""
    s = str(raw_period).strip()
    if not s:
        return None, None, "", ""
    if re.match(r"^\d{5}$", s):
        s = "0" + s
    if re.match(r"^\d{6}$", s):
        month = int(s[:2])
        year = int(s[2:])
        if 1 <= month <= 12:
            return month, year, f"{MONTH_NAMES[month - 1]}-{year}", s
    return parse_gstr1_tax_period(raw_period, financial_year)


def parse_gstr1_tax_period(raw_period: object, financial_year: str = "") -> Tuple[Optional[int], Optional[int], str, str]:
    if raw_period is None:
        return None, None, "", ""
    s = str(raw_period).strip()
    if not s:
        return None, None, "", ""
    if re.match(r"^\d{5,6}$", s):
        return parse_gstr2a_tax_period(s, financial_year)
    clean_key = re.sub(r"[^a-z0-9]", "", s.lower())
    month = MONTH_NAME_MAP.get(clean_key)
    if month:
        month_name = MONTH_NAMES[month - 1]
        start_year, end_year, _ = parse_financial_year(financial_year)
        year = None
        if start_year and end_year:
            year = start_year if month >= 4 else end_year
        display = f"{month_name}-{year}" if year else month_name
        return month, year, display, s
    m = re.match(r"^([A-Za-z]+)[-\s](\d{4})$", s)
    if m:
        m_key = m.group(1).lower()
        if m_key in MONTH_NAME_MAP:
            m_idx = MONTH_NAME_MAP[m_key]
            y = int(m.group(2))
            return m_idx, y, f"{MONTH_NAMES[m_idx - 1]}-{y}", s
    return None, None, s, s


def build_tax_period_range_display(period_objects: List[Dict[str, Any]]) -> str:
    valid = [p for p in period_objects if p.get("month") and p.get("year")]
    if not valid:
        raw_displays = [p.get("display") or p.get("raw") for p in period_objects if p.get("display") or p.get("raw")]
        return raw_displays[0] if raw_displays else ""
    sorted_periods = sorted(valid, key=lambda p: (p["year"], p["month"]))
    first = sorted_periods[0]
    last = sorted_periods[-1]
    if first["year"] == last["year"] and first["month"] == last["month"]:
        return first["display"]
    return f"{first['display']} to {last['display']}"


def _value_to_the_right(grid: List[List[Any]], row_idx: int, start_col: int, max_col: int) -> str:
    if row_idx >= len(grid):
        return ""
    row = grid[row_idx]
    for col in range(start_col + 1, min(max_col + 1, len(row))):
        raw = row[col]
        if raw is None:
            continue
        text = _clean_value(raw)
        if not text:
            continue
        if _match_field(text):
            break
        return text
    return ""


def scan_readme_labels(grid: List[List[Any]], max_row: int = 20, max_col: int = 8) -> Dict[str, str]:
    """Dynamically map metadata fields by scanning label cells in grid."""
    found: Dict[str, str] = {}
    scan_rows = min(max_row, len(grid))

    for r_idx in range(scan_rows):
        row = grid[r_idx]
        scan_cols = min(max_col, len(row))
        for c_idx in range(scan_cols):
            label_cell = row[c_idx]
            field = _match_field(label_cell)
            if not field or field in found:
                continue
            value = _value_to_the_right(grid, r_idx, c_idx, scan_cols)
            if value:
                found[field] = value
    return found


def apply_fallback(
    grid: List[List[Any]],
    found: Dict[str, str],
    fallback: Dict[str, Tuple[int, int]],
) -> Dict[str, str]:
    result = dict(found)
    for field, (row, col) in fallback.items():
        if field in result and result[field]:
            continue
        if row <= 0 or col <= 0:
            continue
        r_idx = row - 1
        c_idx = col - 1
        if r_idx < len(grid) and c_idx < len(grid[r_idx]):
            value = _clean_value(grid[r_idx][c_idx])
            if value:
                result[field] = value
    return result


def fields_to_dealer(fields: Dict[str, str], return_type: str = "gstr2a") -> DealerMetadata:
    _, _, norm_fy = parse_financial_year(fields.get("financial_year", ""))
    raw_period = fields.get("tax_period", "")
    if return_type == "gstr1":
        month, year, disp_period, raw_p = parse_gstr1_tax_period(raw_period, norm_fy)
    else:
        month, year, disp_period, raw_p = parse_gstr2a_tax_period(raw_period, norm_fy)

    return DealerMetadata(
        gstin=fields.get("gstin", "").upper().strip(),
        legal_name=fields.get("legal_name", "").strip(),
        trade_name=fields.get("trade_name", "").strip(),
        financial_year=norm_fy or fields.get("financial_year", "").strip(),
        tax_period=disp_period or raw_period.strip(),
        arn=fields.get("arn", "").strip(),
        arn_date=fields.get("arn_date", "").strip(),
        download_date=fields.get("download_date", "").strip(),
    ).ensure_id()


def extract_from_workbook(wb, return_type: str) -> DealerMetadata:
    readme_name = find_readme_sheet_name(wb.sheetnames)
    if not readme_name:
        return DealerMetadata().ensure_id()

    ws = wb[readme_name]
    raw_grid = [list(r) for r in ws.iter_rows(values_only=True)]
    scanned = scan_readme_labels(raw_grid)
    fallback = GSTR1_FALLBACK if return_type == "gstr1" else GSTR2A_FALLBACK
    merged = apply_fallback(raw_grid, scanned, fallback)
    return fields_to_dealer(merged, return_type)


def extract_from_bytes(content: bytes, return_type: str) -> DealerMetadata:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        return extract_from_workbook(wb, return_type)
    finally:
        wb.close()


def extract_from_files(
    files: List[Tuple[str, bytes]],
    return_type: str,
) -> List[Tuple[str, DealerMetadata]]:
    if return_type not in {"gstr1", "gstr2a"}:
        raise ValueError(f"Unsupported return type: {return_type}")

    results: List[Tuple[str, DealerMetadata]] = []
    for filename, content in files:
        dealer = extract_from_bytes(content, return_type)
        results.append((filename, dealer))
    return results
