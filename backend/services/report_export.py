"""Generate audit report exports with dealer metadata cover page."""

from __future__ import annotations

import io
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from models.dealer_metadata import DealerMetadata


def _dealer_rows(dealer: DealerMetadata, current_dataset: str) -> list[tuple[str, str]]:
    return [
        ("Legal Name", dealer.legal_name or "—"),
        ("Trade Name", dealer.trade_name or "—"),
        ("GSTIN", dealer.gstin or "—"),
        ("Financial Year", dealer.financial_year or "—"),
        ("Tax Period", dealer.tax_period or "—"),
        ("ARN", dealer.arn or "—"),
        ("ARN Date", dealer.arn_date or "—"),
        ("Download Date", dealer.download_date or "—"),
        ("Current Dataset", current_dataset or "—"),
    ]


def build_excel_audit_report(
    dealer: DealerMetadata,
    *,
    current_dataset: str = "",
    report_title: str = "GST Audit Report",
) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dealer Information"

    title_font = Font(bold=True, size=16, color="1E3A8A")
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="DBEAFE")

    ws["A1"] = report_title
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")

    ws["A3"] = "Dealer Metadata"
    ws["A3"].font = Font(bold=True, size=12)

    row = 5
    for label, value in _dealer_rows(dealer, current_dataset):
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        cell = ws.cell(row=row, column=2, value=value)
        cell.alignment = Alignment(wrap_text=True)
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 48

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_pdf_audit_report(
    dealer: DealerMetadata,
    *,
    current_dataset: str = "",
    report_title: str = "GST Audit Report",
) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        title=report_title,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#1E3A8A"),
        spaceAfter=16,
    )
    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontSize=13,
        spaceBefore=8,
        spaceAfter=10,
    )

    story = [
        Paragraph(report_title, title_style),
        Paragraph("Dealer Information", section_style),
        Spacer(1, 0.2 * cm),
    ]

    table_data = [["Field", "Value"]] + [
        [label, value] for label, value in _dealer_rows(dealer, current_dataset)
    ]
    table = Table(table_data, colWidths=[5.5 * cm, 11 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DBEAFE")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(table)

    doc.build(story)
    buffer.seek(0)
    return buffer


def safe_report_filename(
    dealer: DealerMetadata,
    extension: str,
    prefix: str = "GST_Audit_Report",
) -> str:
    gstin = dealer.gstin or "UNKNOWN"
    fy = dealer.financial_year or "UNKNOWN"
    safe = lambda s: "".join(c if c.isalnum() or c in "-_" else "_" for c in s)
    return f"{prefix}_{safe(gstin)}_{safe(fy)}.{extension}"
