import io
import re
import copy as _copy_module
from datetime import datetime
from typing import List, Tuple, Dict, Any, Optional, Set
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font as OxlFont

from models.dealer_metadata import DealerMetadata, WorkbookMetadataResponse
from services.dealer_metadata_service import extract_from_files
from services.dealer_validation import DealerValidationError, validate_dealer_consistency

MONTH_MAP = {
    '01': 'Jan', '02': 'Feb', '03': 'Mar', '04': 'Apr',
    '05': 'May', '06': 'Jun', '07': 'Jul', '08': 'Aug',
    '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dec',
}

FULL_MONTH_MAP = {
    '01': 'January',  '02': 'February', '03': 'March',    '04': 'April',
    '05': 'May',      '06': 'June',     '07': 'July',     '08': 'August',
    '09': 'September','10': 'October',  '11': 'November', '12': 'December',
}

GSTR1_SKIP_SHEETS = {'read me', 'readme', 'gst_audit_meta'}
GSTR2A_SKIP_SHEETS = {'read me', 'readme', 'gst_audit_meta'}
GSTR1_HEADER_ROW = 3

README_TAX_PERIOD_ROW = 4
README_ARN_ROW = 8
README_ARN_DATE_ROW = 9
README_VALUE_COL = 2

GSTIN_PATTERN = re.compile(r'^\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]Z[A-Z0-9]$')
DATE_PATTERN = re.compile(r'^\d{2}-\d{2}-\d{4}$')

GSTR2A_HEADER_HINTS = (
    'gstin', 'invoice number', 'document number', 'note number',
    'original details', 'revised details', 'trade/legal', 'place of supply',
    'taxable value', 'document type', 'invoice type', 'note type',
    'eligibility of itc', 'isd document', 'document details',
)

# ---- Financial-year helpers ----
def fy_sort_key(mm: str, yyyy: str) -> int:
    mm_int, yyyy_int = int(mm), int(yyyy)
    if mm_int >= 4:
        return yyyy_int * 100 + (mm_int - 3)
    else:
        return (yyyy_int - 1) * 100 + (mm_int + 9)

def fy_key_to_label(key: int) -> str:
    idx = key % 100
    fy_start_yr = key // 100
    if idx <= 9:
        cal_month = idx + 3
        cal_year = fy_start_yr
    else:
        cal_month = idx - 9
        cal_year = fy_start_yr + 1
    return f"{FULL_MONTH_MAP.get(f'{cal_month:02d}', str(cal_month))} {cal_year}"

def next_fy_key(key: int) -> int:
    idx = key % 100
    fy_year = key // 100
    if idx < 12:
        return fy_year * 100 + (idx + 1)
    return (fy_year + 1) * 100 + 1

def file_fy_key(filename: str) -> int:
    m = re.search(r'_(\d{2})(\d{4})_', filename)
    if m:
        return fy_sort_key(m.group(1), m.group(2))
    return 999999

def extract_period(filename: str) -> str:
    m = re.search(r'_(\d{2})(\d{4})_', filename)
    if m:
        mm, yyyy = m.group(1), m.group(2)
        return f"{MONTH_MAP.get(mm, mm)}-{yyyy}"
    name_without_ext = filename.split('.')[0]
    return name_without_ext

def find_missing_months(filenames: List[str]) -> List[str]:
    present = set()
    for fname in filenames:
        m = re.search(r'_(\d{2})(\d{4})_', fname)
        if m:
            present.add(fy_sort_key(m.group(1), m.group(2)))

    if len(present) < 2:
        return []

    min_key = min(present)
    max_key = max(present)
    missing = []
    cur = next_fy_key(min_key)
    while cur < max_key:
        if cur not in present:
            missing.append(fy_key_to_label(cur))
        cur = next_fy_key(cur)
    return missing

# ---- Style helpers ----
def copy_cell_style(src, dst):
    if src is None or not src.has_style:
        return
    dst.font          = _copy_module.copy(src.font)
    dst.fill          = _copy_module.copy(src.fill)
    dst.border        = _copy_module.copy(src.border)
    dst.alignment     = _copy_module.copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection    = _copy_module.copy(src.protection)

def apply_data_style(src_hdr, dst):
    if src_hdr is not None and src_hdr.has_style:
        dst.number_format = src_hdr.number_format
    dst.fill = PatternFill(fill_type=None)
    dst.font = OxlFont(color='FF000000')

def safe_value(val):
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(val, (np.integer, int)):
        return int(val)
    if isinstance(val, (np.floating, float)):
        return float(val)
    if isinstance(val, (np.bool_, bool)):
        return bool(val)
    return val

def clean_str(val: Any) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    return str(val).strip()

def normalize_rate(val: Any) -> Optional[float]:
    """
    Normalizes GST Rate values for comparison and grouping.
    Crucially distinguishes 0.0 (0% GST rate) from None (missing/'-'/NA).
    """
    if val is None:
        return None
    s = clean_str(val)
    if not s or s in ('-', '—') or s.lower() in ('na', 'n/a'):
        return None
    clean_num = s.replace('%', '').strip()
    try:
        num = float(clean_num)
        return round(num, 4)
    except (ValueError, TypeError):
        return None

def format_clean_rate(val: Any) -> Any:
    norm = normalize_rate(val)
    if norm is None:
        return '-'
    if norm == int(norm):
        return float(norm)
    return norm

def normalize_numeric(val: Any) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return 0.0 if pd.isna(val) else float(val)
    s = clean_str(val).replace(',', '').replace('₹', '').replace('$', '').strip()
    if not s or s in ('-', '—'):
        return 0.0
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0

def is_portal_total_row(row_values: List[Any], num_col_idx: int = -1) -> bool:
    """
    Structural detection of portal-generated Total / Subtotal rows.
    """
    if not row_values:
        return False
    if 0 <= num_col_idx < len(row_values):
        val = clean_str(row_values[num_col_idx])
        if re.search(r'-(?:total|subtotal)$', val, re.I):
            return True
        if re.match(r'^(?:total|grand total|subtotal)$', val, re.I):
            return True

    for c, v in enumerate(row_values):
        s = clean_str(v)
        if not s:
            continue
        if re.search(r'-(?:total|subtotal)$', s, re.I):
            return True
        if c <= 5 and re.match(r'^(?:total|grand total|subtotal)$', s, re.I):
            return True

    return False

# ---- E-Way Bill Merger ----
def find_eway_date_column(df: pd.DataFrame):
    """Locate the 'EWB No & Dt' column (typically column F)."""
    for col in df.columns:
        col_text = str(col).lower().replace('\n', ' ').replace('  ', ' ')
        if 'ewb' in col_text and ('dt' in col_text or 'date' in col_text):
            return col
    if len(df.columns) > 5:
        return df.columns[5]
    return None

def parse_eway_source_period(value) -> str:
    """Parse month-year from EWB cell like '101581579034 - 11/01/2023 10:29:00'."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ''

    if hasattr(value, 'month') and hasattr(value, 'year'):
        return f"{MONTH_MAP.get(f'{value.month:02d}', value.month)}-{value.year}"

    text = str(value).strip()
    if not text:
        return ''

    date_text = text.split(' - ', 1)[-1].strip() if ' - ' in text else text
    date_text = date_text.split()[0] if ' ' in date_text else date_text

    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y'):
        try:
            parsed = datetime.strptime(date_text, fmt)
            return f"{MONTH_MAP.get(f'{parsed.month:02d}', parsed.month)}-{parsed.year}"
        except ValueError:
            continue

    return ''

def merge_eway_bills(files: List[Tuple[str, bytes]]) -> io.BytesIO:
    if not files:
        raise ValueError("No files provided for merging.")

    all_dfs = []
    for filename, content in files:
        try:
            if filename.lower().endswith('.xls'):
                header = content[:500].lower()
                if b'<html' in header or b'<!doctype html' in header or b'<style' in header:
                    tables = pd.read_html(io.BytesIO(content))
                    if not tables:
                        raise ValueError(f"No tables found in HTML-xls file {filename}")
                    df_temp = tables[0]
                    xlsx_buffer = io.BytesIO()
                    df_temp.to_excel(xlsx_buffer, index=False, engine='openpyxl')
                    xlsx_buffer.seek(0)
                    xls = pd.ExcelFile(xlsx_buffer, engine='openpyxl')
                else:
                    xls = pd.ExcelFile(io.BytesIO(content), engine='xlrd')
            else:
                xls = pd.ExcelFile(io.BytesIO(content), engine='openpyxl')

            for sheet_name in xls.sheet_names:
                try:
                    df = pd.read_excel(xls, sheet_name=sheet_name, engine='openpyxl')
                except Exception:
                    df = pd.read_excel(xls, sheet_name=sheet_name)

                date_col = find_eway_date_column(df)
                if date_col is not None:
                    df['Source_Period'] = df[date_col].apply(parse_eway_source_period)
                else:
                    df['Source_Period'] = ''

                all_dfs.append(df)
        except Exception as e:
            raise ValueError(f"Error processing {filename}: {str(e)}")

    if not all_dfs:
        raise ValueError("No data found to merge.")

    combined_df = pd.concat(all_dfs, ignore_index=True)
    output_buffer = io.BytesIO()
    combined_df.to_excel(output_buffer, index=False, engine='openpyxl')
    output_buffer.seek(0)
    return output_buffer

# ---- GSTR-1 Merger ----
def build_gstr1_col_map(headers: List[str]) -> Dict[str, int]:
    col_map = {
        'gstin': -1, 'receiver_name': -1, 'inv_number': -1, 'inv_date': -1,
        'inv_value': -1, 'pos': -1, 'reverse_charge': -1, 'inv_type': -1,
        'rate': -1, 'taxable_value': -1, 'igst': -1, 'cgst': -1, 'sgst': -1,
        'cess': -1, 'note_number': -1, 'note_date': -1, 'note_type': -1,
        'note_supply_type': -1, 'note_value': -1, 'hsn': -1, 'description': -1,
        'uqc': -1, 'total_qty': -1, 'doc_nature': -1, 'sr_from': -1, 'sr_to': -1,
        'total_docs': -1, 'cancelled_docs': -1, 'nil_supplies': -1,
        'exempt_supplies': -1, 'non_gst_supplies': -1, 'gross_advance': -1,
    }
    for c, h in enumerate(headers):
        text = clean_str(h).lower()
        if not text:
            continue
        if re.search(r'gstin|uin', text): col_map['gstin'] = c
        if 'receiver name' in text: col_map['receiver_name'] = c
        if re.search(r'invoice number|inv no', text) and 'original' not in text: col_map['inv_number'] = c
        if re.search(r'invoice date', text) and 'original' not in text: col_map['inv_date'] = c
        if 'invoice value' in text: col_map['inv_value'] = c
        if 'place of supply' in text: col_map['pos'] = c
        if 'reverse charge' in text: col_map['reverse_charge'] = c
        if 'invoice type' in text: col_map['inv_type'] = c
        if re.search(r'^rate$|^rate\s*\(|gst rate', text): col_map['rate'] = c
        if re.search(r'taxable value|taxable val', text): col_map['taxable_value'] = c
        if re.search(r'integrated tax|igst', text): col_map['igst'] = c
        if re.search(r'central tax|cgst', text): col_map['cgst'] = c
        if re.search(r'state.*tax|sgst|ut tax', text): col_map['sgst'] = c
        if 'cess' in text: col_map['cess'] = c
        if re.search(r'note number', text) and 'original' not in text: col_map['note_number'] = c
        if re.search(r'note date', text) and 'original' not in text: col_map['note_date'] = c
        if re.search(r'note type', text) and 'original' not in text: col_map['note_type'] = c
        if 'note supply type' in text: col_map['note_supply_type'] = c
        if 'note value' in text: col_map['note_value'] = c
        if re.search(r'^hsn', text): col_map['hsn'] = c
        if 'description' in text: col_map['description'] = c
        if 'uqc' in text: col_map['uqc'] = c
        if 'total quantity' in text: col_map['total_qty'] = c
        if 'nature of document' in text: col_map['doc_nature'] = c
        if re.search(r'sr\.?\s*no\.?\s*from', text): col_map['sr_from'] = c
        if re.search(r'sr\.?\s*no\.?\s*to', text): col_map['sr_to'] = c
        if 'total number' in text: col_map['total_docs'] = c
        if 'cancelled' in text: col_map['cancelled_docs'] = c
        if 'nil rated supplies' in text: col_map['nil_supplies'] = c
        if 'exempted' in text: col_map['exempt_supplies'] = c
        if 'non-gst supplies' in text: col_map['non_gst_supplies'] = c
        if 'gross advance' in text: col_map['gross_advance'] = c
    return col_map

def process_gstr1_sheet_rows(sheet_name: str, raw_rows: List[List[Any]], col_map: Dict[str, int], period: str):
    sheet_lower = sheet_name.strip().lower()
    detail_rows = []
    doc_col = col_map['inv_number'] if col_map['inv_number'] != -1 else col_map['note_number']

    # Operation 1: Filter out portal totals if any
    for row in raw_rows:
        if not any(clean_str(v) != '' for v in row):
            continue
        if is_portal_total_row(row, doc_col):
            continue
        detail_rows.append(row)

    # Operation 2: Within-document detail aggregation
    aggregated_map: Dict[str, List[Any]] = {}

    cur_gstin = ''
    cur_inv_no = ''
    cur_inv_date = ''
    cur_inv_type = ''
    cur_note_no = ''
    cur_note_date = ''
    cur_note_type = ''
    cur_pos = ''
    cur_rev_chg = ''

    for row in detail_rows:
        raw_gstin = clean_str(row[col_map['gstin']]).upper() if col_map['gstin'] != -1 and col_map['gstin'] < len(row) else ''
        raw_inv_no = clean_str(row[col_map['inv_number']]).upper() if col_map['inv_number'] != -1 and col_map['inv_number'] < len(row) else ''
        raw_note_no = clean_str(row[col_map['note_number']]).upper() if col_map['note_number'] != -1 and col_map['note_number'] < len(row) else ''

        if raw_gstin:
            cur_gstin = raw_gstin
        if raw_inv_no:
            cur_inv_no = raw_inv_no
        if raw_note_no:
            cur_note_no = raw_note_no
        if col_map['inv_date'] != -1 and col_map['inv_date'] < len(row) and clean_str(row[col_map['inv_date']]):
            cur_inv_date = clean_str(row[col_map['inv_date']])
        if col_map['inv_type'] != -1 and col_map['inv_type'] < len(row) and clean_str(row[col_map['inv_type']]):
            cur_inv_type = clean_str(row[col_map['inv_type']])
        if col_map['note_date'] != -1 and col_map['note_date'] < len(row) and clean_str(row[col_map['note_date']]):
            cur_note_date = clean_str(row[col_map['note_date']])
        if col_map['note_type'] != -1 and col_map['note_type'] < len(row) and clean_str(row[col_map['note_type']]):
            cur_note_type = clean_str(row[col_map['note_type']])
        if col_map['pos'] != -1 and col_map['pos'] < len(row) and clean_str(row[col_map['pos']]):
            cur_pos = clean_str(row[col_map['pos']])
        if col_map['reverse_charge'] != -1 and col_map['reverse_charge'] < len(row) and clean_str(row[col_map['reverse_charge']]):
            cur_rev_chg = clean_str(row[col_map['reverse_charge']])

        gstin = raw_gstin or cur_gstin
        inv_no = raw_inv_no or cur_inv_no
        inv_date = clean_str(row[col_map['inv_date']]) if col_map['inv_date'] != -1 and col_map['inv_date'] < len(row) and clean_str(row[col_map['inv_date']]) else cur_inv_date
        inv_type = clean_str(row[col_map['inv_type']]) if col_map['inv_type'] != -1 and col_map['inv_type'] < len(row) and clean_str(row[col_map['inv_type']]) else cur_inv_type
        note_no = raw_note_no or cur_note_no
        note_date = clean_str(row[col_map['note_date']]) if col_map['note_date'] != -1 and col_map['note_date'] < len(row) and clean_str(row[col_map['note_date']]) else cur_note_date
        note_type = clean_str(row[col_map['note_type']]) if col_map['note_type'] != -1 and col_map['note_type'] < len(row) and clean_str(row[col_map['note_type']]) else cur_note_type
        pos = clean_str(row[col_map['pos']]) if col_map['pos'] != -1 and col_map['pos'] < len(row) and clean_str(row[col_map['pos']]) else cur_pos
        rev_chg = clean_str(row[col_map['reverse_charge']]) if col_map['reverse_charge'] != -1 and col_map['reverse_charge'] < len(row) and clean_str(row[col_map['reverse_charge']]) else cur_rev_chg

        hsn = clean_str(row[col_map['hsn']]) if col_map['hsn'] != -1 and col_map['hsn'] < len(row) else ''
        uqc = clean_str(row[col_map['uqc']]) if col_map['uqc'] != -1 and col_map['uqc'] < len(row) else ''
        desc = clean_str(row[col_map['description']]) if col_map['description'] != -1 and col_map['description'] < len(row) else ''
        doc_nature = clean_str(row[col_map['doc_nature']]) if col_map['doc_nature'] != -1 and col_map['doc_nature'] < len(row) else ''
        sr_from = clean_str(row[col_map['sr_from']]) if col_map['sr_from'] != -1 and col_map['sr_from'] < len(row) else ''
        sr_to = clean_str(row[col_map['sr_to']]) if col_map['sr_to'] != -1 and col_map['sr_to'] < len(row) else ''

        raw_rate = row[col_map['rate']] if col_map['rate'] != -1 and col_map['rate'] < len(row) else None
        norm_rate = normalize_rate(raw_rate)
        rate_key = 'NORATE' if norm_rate is None else f"R{norm_rate}"

        if any(k in sheet_lower for k in ('b2b', 'b2cl', 'exp')):
            doc_key = f"{gstin}|{inv_no}|{inv_date}|{inv_type}"
            agg_key = f"{doc_key}|{pos}|{rev_chg}|{rate_key}|{period}"
        elif any(k in sheet_lower for k in ('cdnr', 'cdnur')):
            doc_key = f"{gstin}|{note_no}|{note_date}|{note_type}"
            agg_key = f"{doc_key}|{pos}|{rev_chg}|{rate_key}|{period}"
        elif 'b2cs' in sheet_lower:
            doc_key = f"{pos}|{rate_key}"
            agg_key = f"{doc_key}|{period}"
        elif 'exemp' in sheet_lower:
            doc_key = f"{desc}"
            agg_key = f"{doc_key}|{period}"
        elif 'hsn' in sheet_lower:
            doc_key = f"{hsn}|{desc}|{uqc}|{rate_key}"
            agg_key = f"{doc_key}|{period}"
        elif 'docs' in sheet_lower:
            doc_key = f"{doc_nature}|{sr_from}|{sr_to}"
            agg_key = f"{doc_key}|{period}"
        elif 'at' in sheet_lower:
            doc_key = f"{pos}|{rate_key}"
            agg_key = f"{doc_key}|{period}"
        else:
            doc_key = '|'.join(clean_str(c) for c in row)
            agg_key = f"{doc_key}|{period}"

        if agg_key not in aggregated_map:
            cloned = list(row)
            if col_map['gstin'] != -1 and col_map['gstin'] < len(cloned) and gstin:
                cloned[col_map['gstin']] = gstin
            if col_map['inv_number'] != -1 and col_map['inv_number'] < len(cloned) and inv_no:
                cloned[col_map['inv_number']] = inv_no
            if col_map['inv_date'] != -1 and col_map['inv_date'] < len(cloned) and inv_date:
                cloned[col_map['inv_date']] = inv_date
            if col_map['inv_type'] != -1 and col_map['inv_type'] < len(cloned) and inv_type:
                cloned[col_map['inv_type']] = inv_type
            if col_map['note_number'] != -1 and col_map['note_number'] < len(cloned) and note_no:
                cloned[col_map['note_number']] = note_no
            if col_map['note_date'] != -1 and col_map['note_date'] < len(cloned) and note_date:
                cloned[col_map['note_date']] = note_date
            if col_map['note_type'] != -1 and col_map['note_type'] < len(cloned) and note_type:
                cloned[col_map['note_type']] = note_type
            if col_map['pos'] != -1 and col_map['pos'] < len(cloned) and pos:
                cloned[col_map['pos']] = pos
            if col_map['reverse_charge'] != -1 and col_map['reverse_charge'] < len(cloned) and rev_chg:
                cloned[col_map['reverse_charge']] = rev_chg
            if col_map['rate'] != -1 and col_map['rate'] < len(cloned):
                cloned[col_map['rate']] = format_clean_rate(raw_rate)
            aggregated_map[agg_key] = cloned
        else:
            target = aggregated_map[agg_key]
            for field in ('taxable_value', 'igst', 'cgst', 'sgst', 'cess', 'total_qty',
                          'nil_supplies', 'exempt_supplies', 'non_gst_supplies', 'gross_advance'):
                idx = col_map[field]
                if idx != -1 and idx < len(target) and idx < len(row):
                    target[idx] = normalize_numeric(target[idx]) + normalize_numeric(row[idx])

    return detail_rows, list(aggregated_map.values())

def merge_gstr1_files(
    files: List[Tuple[str, bytes]],
) -> Tuple[io.BytesIO, str, List[str], DealerMetadata, str]:
    if not files:
        raise ValueError("No files provided for merging.")

    dealer_records = extract_from_files(files, "gstr1")
    dealer = validate_dealer_consistency(dealer_records)

    filenames = [f[0] for f in files]
    missing = find_missing_months(filenames)
    sorted_files = sorted(files, key=lambda f: file_fy_key(f[0]))

    period_keys = []
    for filename, _ in sorted_files:
        m = re.search(r'_(\d{2})(\d{4})_', filename)
        if m:
            mm, yyyy = m.group(1), m.group(2)
            period_keys.append(
                (int(yyyy) * 100 + int(mm), FULL_MONTH_MAP.get(mm, mm), yyyy)
            )

    if period_keys:
        period_keys.sort()
        first_label = f"{period_keys[0][1]} {period_keys[0][2]}"
        last_label  = f"{period_keys[-1][1]} {period_keys[-1][2]}"
        tax_period_text = (
            first_label if first_label == last_label
            else f"{first_label} to {last_label}"
        )
    else:
        tax_period_text = "Full Period"

    first_filename, first_content = sorted_files[0]
    wb = openpyxl.load_workbook(io.BytesIO(first_content))

    sheet_collected_items: Dict[str, List[Tuple[str, List[Any]]]] = {}

    for filename, content in sorted_files:
        period = extract_period(filename)
        src_wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        for sheet_name in src_wb.sheetnames:
            if sheet_name.strip().lower() in GSTR1_SKIP_SHEETS:
                continue

            ws_src = src_wb[sheet_name]
            raw_rows = [list(row) for row in ws_src.iter_rows(values_only=True)]

            if len(raw_rows) <= GSTR1_HEADER_ROW:
                continue

            header_row = [clean_str(v) for v in raw_rows[GSTR1_HEADER_ROW]]
            col_map = build_gstr1_col_map(header_row)
            data_rows = raw_rows[GSTR1_HEADER_ROW + 1:]

            _, agg_rows = process_gstr1_sheet_rows(sheet_name, data_rows, col_map, period)

            if sheet_name not in sheet_collected_items:
                sheet_collected_items[sheet_name] = []

            for agg_row in agg_rows:
                sheet_collected_items[sheet_name].append((period, agg_row))

        src_wb.close()

    # Step 3b: Auto-generate output filename from Read me metadata
    readme_ws_name = next(
        (sn for sn in wb.sheetnames if sn.strip().lower() == 'read me'),
        None
    )
    val_col = README_VALUE_COL + 1

    if dealer.gstin and dealer.financial_year:
        safe_gstin = re.sub(r'[\\/:*?"<>|]', '_', dealer.gstin)
        safe_fy = re.sub(r'[\\/:*?"<>|]', '_', dealer.financial_year)
        auto_name = f"GSTR1_{safe_gstin}_{safe_fy}_Merged.xlsx"
    elif readme_ws_name:
        ws_rm_meta = wb[readme_ws_name]
        gstin_val = ws_rm_meta.cell(row=6, column=val_col).value or ''
        fy_val = ws_rm_meta.cell(row=4, column=val_col).value or ''
        safe_gstin = re.sub(r'[\\/:*?"<>|]', '_', str(gstin_val).strip())
        safe_fy = re.sub(r'[\\/:*?"<>|]', '_', str(fy_val).strip())
        auto_name = f"GSTR1_{safe_gstin}_{safe_fy}_Merged.xlsx"
    else:
        auto_name = "GSTR1_Merged.xlsx"

    if readme_ws_name:
        ws_rm = wb[readme_ws_name]
        ws_rm.cell(row=README_TAX_PERIOD_ROW + 1, column=val_col).value = tax_period_text
        ws_rm.cell(row=README_ARN_ROW      + 1, column=val_col).value = ''
        ws_rm.cell(row=README_ARN_DATE_ROW + 1, column=val_col).value = ''

    HEADER_ROW = GSTR1_HEADER_ROW + 1
    DATA_START = HEADER_ROW + 1

    # Operation 3: Cross-file duplicate detection and final workbook build
    seen_keys: Set[str] = set()

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in GSTR1_SKIP_SHEETS:
            continue
        if sheet_name not in sheet_collected_items:
            continue

        ws = wb[sheet_name]
        max_data_col = ws.max_column
        hdr_cells = {cell.column: cell for cell in ws[HEADER_ROW]}

        if ws.max_row >= DATA_START:
            ws.delete_rows(DATA_START, ws.max_row - DATA_START + 1)

        sp_col = max_data_col + 1
        sp_ref_hdr = hdr_cells.get(max_data_col)
        sp_hdr_cell = ws.cell(row=HEADER_ROW, column=sp_col, value='Source_Period')
        copy_cell_style(sp_ref_hdr, sp_hdr_cell)

        write_row_idx = DATA_START
        for period, row_vals in sheet_collected_items[sheet_name]:
            # Cross file duplicate check key
            row_clean = [clean_str(v).upper() for v in row_vals]
            dup_key = f"{sheet_name}|{period}|{'|'.join(row_clean)}"
            if dup_key in seen_keys:
                continue
            seen_keys.add(dup_key)

            for c_idx, val in enumerate(row_vals):
                col_num = c_idx + 1
                dst_cell = ws.cell(row=write_row_idx, column=col_num, value=safe_value(val))
                apply_data_style(hdr_cells.get(col_num), dst_cell)

            sp_cell = ws.cell(row=write_row_idx, column=sp_col, value=period)
            apply_data_style(sp_ref_hdr, sp_cell)
            write_row_idx += 1

    dealer.tax_period = tax_period_text
    workbook_id = WorkbookMetadataResponse.build_workbook_id("gstr1", dealer, filenames)

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer, auto_name, missing, dealer, workbook_id

# ---- GSTR-2A Merger ----
def build_gstr2a_col_map(header_block: List[List[Any]]) -> Dict[str, int]:
    col_map = {
        'gstin': -1, 'trade_name': -1, 'inv_number': -1, 'inv_type': -1,
        'inv_date': -1, 'inv_value': -1, 'pos': -1, 'reverse_charge': -1,
        'rate': -1, 'taxable_value': -1, 'igst': -1, 'cgst': -1, 'sgst': -1,
        'cess': -1, 'note_number': -1, 'note_type': -1, 'note_date': -1,
    }
    num_cols = max((len(r) for r in header_block), default=0)
    for c in range(num_cols):
        cell_texts = [clean_str(r[c]).lower() for r in header_block if c < len(r) and clean_str(r[c])]
        joined = ' '.join(cell_texts)

        if 'gstin' in joined and col_map['gstin'] == -1: col_map['gstin'] = c
        if re.search(r'trade|legal name', joined) and col_map['trade_name'] == -1: col_map['trade_name'] = c
        if re.search(r'invoice number|inv no', joined) and 'original' not in joined:
            if col_map['inv_number'] == -1 or 'details' in joined: col_map['inv_number'] = c
        if 'invoice type' in joined and col_map['inv_type'] == -1: col_map['inv_type'] = c
        if 'invoice date' in joined and 'original' not in joined and col_map['inv_date'] == -1: col_map['inv_date'] = c
        if re.search(r'invoice value|note value|document value', joined) and col_map['inv_value'] == -1: col_map['inv_value'] = c
        if 'place of supply' in joined and col_map['pos'] == -1: col_map['pos'] = c
        if 'reverse charge' in joined and col_map['reverse_charge'] == -1: col_map['reverse_charge'] = c
        if re.search(r'^rate|^rate\s*\(|gst rate', joined) or any(t.startswith('rate') for t in cell_texts):
            if col_map['rate'] == -1: col_map['rate'] = c
        if 'taxable value' in joined and col_map['taxable_value'] == -1: col_map['taxable_value'] = c
        if re.search(r'integrated tax|igst', joined) and col_map['igst'] == -1: col_map['igst'] = c
        if re.search(r'central tax|cgst', joined) and col_map['cgst'] == -1: col_map['cgst'] = c
        if re.search(r'state.*tax|sgst|ut tax', joined) and col_map['sgst'] == -1: col_map['sgst'] = c
        if 'cess' in joined and col_map['cess'] == -1: col_map['cess'] = c
        if 'note number' in joined and 'original' not in joined and col_map['note_number'] == -1: col_map['note_number'] = c
        if 'note type' in joined and 'original' not in joined and col_map['note_type'] == -1: col_map['note_type'] = c
        if 'note date' in joined and 'original' not in joined and col_map['note_date'] == -1: col_map['note_date'] = c

    return col_map

def is_gstr2a_data_row(row_values: List[Any]) -> bool:
    if not any(clean_str(v) != '' for v in row_values):
        return False
    row_text = ' '.join(clean_str(v).lower() for v in row_values)

    if any(hint in row_text for hint in GSTR2A_HEADER_HINTS):
        has_gstin_or_total = any(
            GSTIN_PATTERN.match(clean_str(c)) or clean_str(c).endswith('-Total')
            for c in row_values
        )
        if not has_gstin_or_total:
            return False

    for c, val in enumerate(row_values):
        text = clean_str(val)
        if not text:
            continue
        if GSTIN_PATTERN.match(text):
            return True
        if text.endswith('-Total'):
            return True
        if DATE_PATTERN.match(text):
            return True
        if isinstance(val, (int, float)) and c >= 5:
            return True

    return False

def find_gstr2a_header_end(ws) -> int:
    last_header = 4
    for r in range(5, min(ws.max_row, 12) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(clean_str(v) != '' for v in row_vals):
            continue
        if is_gstr2a_data_row(row_vals):
            break
        last_header = r
    return last_header

def unmerge_from_row(ws, start_row: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row >= start_row:
            ws.unmerge_cells(str(merged_range))

def merge_gstr2a_files(
    files: List[Tuple[str, bytes]],
) -> Tuple[io.BytesIO, str, List[str], DealerMetadata, str]:
    if not files:
        raise ValueError("No files provided for merging.")

    dealer_records = extract_from_files(files, "gstr2a")
    dealer = validate_dealer_consistency(dealer_records)

    filenames = [f[0] for f in files]
    missing = find_missing_months(filenames)
    sorted_files = sorted(files, key=lambda f: file_fy_key(f[0]))

    period_keys = []
    for filename, _ in sorted_files:
        m = re.search(r'_(\d{2})(\d{4})_', filename)
        if m:
            mm, yyyy = m.group(1), m.group(2)
            period_keys.append(
                (int(yyyy) * 100 + int(mm), FULL_MONTH_MAP.get(mm, mm), yyyy)
            )

    if period_keys:
        period_keys.sort()
        first_label = f"{period_keys[0][1]} {period_keys[0][2]}"
        last_label = f"{period_keys[-1][1]} {period_keys[-1][2]}"
        tax_period_text = (
            first_label if first_label == last_label
            else f"{first_label} to {last_label}"
        )
    else:
        tax_period_text = "Full Period"

    first_filename, first_content = sorted_files[0]
    wb = openpyxl.load_workbook(io.BytesIO(first_content))

    sheet_collected_items: Dict[str, List[Tuple[str, List[Any]]]] = {}

    for filename, content in sorted_files:
        period = extract_period(filename)
        src_wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        for sheet_name in src_wb.sheetnames:
            if sheet_name.strip().lower() in GSTR2A_SKIP_SHEETS:
                continue

            ws_src = src_wb[sheet_name]
            all_sheet_rows = [list(r) for r in ws_src.iter_rows(values_only=True)]
            if not all_sheet_rows:
                continue

            header_end = find_gstr2a_header_end(ws_src)
            header_block = all_sheet_rows[:header_end]

            col_map = build_gstr2a_col_map(header_block)
            doc_col = col_map['inv_number'] if col_map['inv_number'] != -1 else col_map['note_number']

            # Operation 1: Remove portal total rows & collect detail rows
            detail_rows = []
            for row_vals in all_sheet_rows[header_end:]:
                if not any(clean_str(v) != '' for v in row_vals):
                    continue
                if not is_gstr2a_data_row(row_vals):
                    continue
                if is_portal_total_row(row_vals, doc_col):
                    continue
                detail_rows.append(row_vals)

            # Operation 2: Within-document detail aggregation
            aggregated_map: Dict[str, List[Any]] = {}
            for row in detail_rows:
                gstin = clean_str(row[col_map['gstin']]).upper() if col_map['gstin'] != -1 and col_map['gstin'] < len(row) else ''
                inv_no = clean_str(row[col_map['inv_number']]).upper() if col_map['inv_number'] != -1 and col_map['inv_number'] < len(row) else ''
                inv_date = clean_str(row[col_map['inv_date']]) if col_map['inv_date'] != -1 and col_map['inv_date'] < len(row) else ''
                inv_type = clean_str(row[col_map['inv_type']]) if col_map['inv_type'] != -1 and col_map['inv_type'] < len(row) else ''
                note_no = clean_str(row[col_map['note_number']]).upper() if col_map['note_number'] != -1 and col_map['note_number'] < len(row) else ''
                note_type = clean_str(row[col_map['note_type']]) if col_map['note_type'] != -1 and col_map['note_type'] < len(row) else ''
                pos = clean_str(row[col_map['pos']]) if col_map['pos'] != -1 and col_map['pos'] < len(row) else ''
                rev_chg = clean_str(row[col_map['reverse_charge']]) if col_map['reverse_charge'] != -1 and col_map['reverse_charge'] < len(row) else ''

                raw_rate = row[col_map['rate']] if col_map['rate'] != -1 and col_map['rate'] < len(row) else None
                norm_rate = normalize_rate(raw_rate)
                rate_key = 'NORATE' if norm_rate is None else f"R{norm_rate}"

                doc_key = f"{gstin}|{inv_no or note_no}|{inv_date or note_type}|{inv_type}"
                agg_key = f"{doc_key}|{pos}|{rev_chg}|{rate_key}|{period}"

                if agg_key not in aggregated_map:
                    cloned = list(row)
                    if col_map['rate'] != -1 and col_map['rate'] < len(cloned):
                        cloned[col_map['rate']] = format_clean_rate(raw_rate)
                    aggregated_map[agg_key] = cloned
                else:
                    target = aggregated_map[agg_key]
                    for field in ('taxable_value', 'igst', 'cgst', 'sgst', 'cess'):
                        idx = col_map[field]
                        if idx != -1 and idx < len(target) and idx < len(row):
                            target[idx] = normalize_numeric(target[idx]) + normalize_numeric(row[idx])

            if sheet_name not in sheet_collected_items:
                sheet_collected_items[sheet_name] = []

            for agg_row in aggregated_map.values():
                sheet_collected_items[sheet_name].append((period, agg_row))

        src_wb.close()

    readme_ws_name = next(
        (sn for sn in wb.sheetnames if sn.strip().lower() == 'read me'),
        None
    )

    if dealer.gstin and dealer.financial_year:
        safe_gstin = re.sub(r'[\\/:*?"<>|]', '_', dealer.gstin)
        safe_fy = re.sub(r'[\\/:*?"<>|]', '_', dealer.financial_year)
        auto_name = f"GSTR2A_{safe_gstin}_{safe_fy}_Merged.xlsx"
    elif readme_ws_name:
        ws_rm_meta = wb[readme_ws_name]
        gstin_val = ws_rm_meta.cell(row=2, column=3).value or ''
        fy_val = ws_rm_meta.cell(row=3, column=5).value or ''
        safe_gstin = re.sub(r'[\\/:*?"<>|]', '_', str(gstin_val).strip())
        safe_fy = re.sub(r'[\\/:*?"<>|]', '_', str(fy_val).strip())
        auto_name = f"GSTR2A_{safe_gstin}_{safe_fy}_Merged.xlsx"
    else:
        auto_name = "GSTR2A_Merged.xlsx"

    if readme_ws_name:
        ws_rm = wb[readme_ws_name]
        ws_rm.cell(row=2, column=5).value = tax_period_text

    # Operation 3: Cross-file duplicate detection and final workbook build
    seen_keys: Set[str] = set()

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in GSTR2A_SKIP_SHEETS:
            continue
        if sheet_name not in sheet_collected_items:
            continue

        ws = wb[sheet_name]
        header_end = find_gstr2a_header_end(ws)
        WRITE_START = header_end + 1
        max_data_col = ws.max_column

        sp_col = max_data_col + 1
        sp_ref_hdr = ws.cell(row=header_end, column=max_data_col)
        if sp_ref_hdr.value in (None, ''):
            sp_ref_hdr = ws.cell(row=max(1, header_end - 1), column=max_data_col)

        sp_hdr_cell = ws.cell(row=header_end, column=sp_col, value='Source_Period')
        copy_cell_style(sp_ref_hdr, sp_hdr_cell)

        unmerge_from_row(ws, WRITE_START)

        if ws.max_row >= WRITE_START:
            ws.delete_rows(WRITE_START, ws.max_row - WRITE_START + 1)

        write_row_idx = WRITE_START
        for period, row_vals in sheet_collected_items[sheet_name]:
            row_clean = [clean_str(v).upper() for v in row_vals]
            dup_key = f"{sheet_name}|{period}|{'|'.join(row_clean)}"
            if dup_key in seen_keys:
                continue
            seen_keys.add(dup_key)

            for c_idx, val in enumerate(row_vals):
                col_num = c_idx + 1
                dst_cell = ws.cell(row=write_row_idx, column=col_num, value=safe_value(val))
                apply_data_style(ws.cell(row=header_end, column=col_num), dst_cell)

            sp_cell = ws.cell(row=write_row_idx, column=sp_col, value=period)
            apply_data_style(sp_ref_hdr, sp_cell)
            write_row_idx += 1

    dealer.tax_period = tax_period_text
    workbook_id = WorkbookMetadataResponse.build_workbook_id("gstr2a", dealer, filenames)

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer, auto_name, missing, dealer, workbook_id
