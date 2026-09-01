import io

import openpyxl
from reportlab.pdfgen import canvas

from models.dealer_metadata import DealerMetadata
from services.report_export import (
    build_excel_audit_report,
    build_pdf_audit_report,
    safe_report_filename,
)


def _sample_dealer() -> DealerMetadata:
    return DealerMetadata(
        gstin="03AABCU9603R1ZX",
        legal_name="UJJIVAN SMALL FINANCE BANK LIMITED",
        trade_name="UJJIVAN SMALL FINANCE BANK LIMITED",
        financial_year="2022-23",
        tax_period="April 2022 to March 2023",
        arn="AA030422181610R",
        arn_date="11-05-2022",
        download_date="08-07-2026",
    ).ensure_id()


class TestReportExport:
    def test_excel_report_contains_dealer_fields(self):
        dealer = _sample_dealer()
        buffer = build_excel_audit_report(dealer, current_dataset="GSTR2A_Merged.xlsx")
        wb = openpyxl.load_workbook(buffer)
        ws = wb.active
        values = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(5, 15)}
        assert values["Legal Name"] == dealer.legal_name
        assert values["GSTIN"] == dealer.gstin
        assert values["Financial Year"] == dealer.financial_year
        assert values["Current Dataset"] == "GSTR2A_Merged.xlsx"

    def test_pdf_report_is_valid_pdf(self):
        dealer = _sample_dealer()
        buffer = build_pdf_audit_report(dealer, current_dataset="GSTR2A_Merged.xlsx")
        data = buffer.getvalue()
        assert data.startswith(b"%PDF")

    def test_safe_report_filename(self):
        dealer = _sample_dealer()
        name = safe_report_filename(dealer, "pdf")
        assert name.endswith(".pdf")
        assert "03AABCU9603R1ZX" in name
