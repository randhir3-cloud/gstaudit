import io

import openpyxl
import pytest

from models.dealer_metadata import DealerMetadata
from services.dealer_metadata_service import extract_from_bytes, scan_readme_labels
from services.dealer_validation import DealerValidationError, validate_dealer_consistency


def _build_gstr1_workbook(
    gstin: str = "03AABCU9603R1ZX",
    fy: str = "2022-23",
    tax_period: str = "April",
    legal_name: str = "UJJIVAN SMALL FINANCE BANK LIMITED",
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Read me"
    ws["A4"] = "Financial Year"
    ws["C4"] = fy
    ws["A5"] = "Tax Period"
    ws["C5"] = tax_period
    ws["A6"] = "GSTIN"
    ws["C6"] = gstin
    ws["A7"] = "Legal Name"
    ws["C7"] = legal_name
    ws["A8"] = "Trade Name (if any)"
    ws["C8"] = legal_name
    ws["A9"] = "ARN"
    ws["C9"] = "AA030422181610R"
    ws["A10"] = "ARN date"
    ws["C10"] = "11-05-2022"
    ws["A11"] = "Date and Time of Generation"
    ws["C11"] = "08/07/2026 11:31:AM"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_gstr2a_workbook(
    gstin: str = "03AABCU9603R1ZX",
    fy: str = "2022-23",
    tax_period: str = "April 2022 to March 2023",
    legal_name: str = "UJJIVAN SMALL FINANCE BANK LIMITED",
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Read me"
    ws["B2"] = "Taxpayer's GSTIN"
    ws["C2"] = gstin
    ws["D2"] = "Tax period"
    ws["E2"] = tax_period
    ws["B3"] = "Legal name"
    ws["C3"] = legal_name
    ws["D3"] = "Financial year"
    ws["E3"] = fy
    ws["B4"] = "Trade name"
    ws["C4"] = legal_name
    ws["D4"] = "Date of generation"
    ws["E4"] = "08-07-2026"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestDealerMetadataExtraction:
    def test_extract_gstr1_by_label_scan(self):
        content = _build_gstr1_workbook()
        dealer = extract_from_bytes(content, "gstr1")
        assert dealer.gstin == "03AABCU9603R1ZX"
        assert dealer.legal_name == "UJJIVAN SMALL FINANCE BANK LIMITED"
        assert dealer.financial_year == "2022-23"
        assert dealer.tax_period == "April"
        assert dealer.arn == "AA030422181610R"
        assert dealer.id is not None

    def test_extract_gstr2a_by_label_scan(self):
        content = _build_gstr2a_workbook()
        dealer = extract_from_bytes(content, "gstr2a")
        assert dealer.gstin == "03AABCU9603R1ZX"
        assert dealer.financial_year == "2022-23"
        assert dealer.tax_period == "April 2022 to March 2023"
        assert dealer.download_date == "08-07-2026"

    def test_extract_gstr1_fallback_positions(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Read me"
        ws["C4"] = "2021-22"
        ws["C5"] = "March"
        ws["C6"] = "29AABCT1332L000"
        ws["C7"] = "Fallback Legal Name"
        buffer = io.BytesIO()
        wb.save(buffer)

        dealer = extract_from_bytes(buffer.getvalue(), "gstr1")
        assert dealer.gstin == "29AABCT1332L000"
        assert dealer.financial_year == "2021-22"
        assert dealer.legal_name == "Fallback Legal Name"

    def test_scan_readme_labels_handles_trade_name_alias(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A8"] = "Trade Name (if any)"
        ws["C8"] = "ACME TRADERS"
        fields = scan_readme_labels(ws)
        assert fields["trade_name"] == "ACME TRADERS"


class TestDealerValidation:
    def test_validate_consistent_files(self):
        records = [
            ("a.xlsx", DealerMetadata(gstin="03AABCU9603R1ZX", financial_year="2022-23")),
            ("b.xlsx", DealerMetadata(gstin="03AABCU9603R1ZX", financial_year="2022-23")),
        ]
        dealer = validate_dealer_consistency(records)
        assert dealer.gstin == "03AABCU9603R1ZX"

    def test_reject_gstin_mismatch(self):
        records = [
            ("a.xlsx", DealerMetadata(gstin="03AABCU9603R1ZX", financial_year="2022-23")),
            ("b.xlsx", DealerMetadata(gstin="29AABCT1332L000", financial_year="2022-23")),
        ]
        with pytest.raises(DealerValidationError) as exc:
            validate_dealer_consistency(records)
        assert exc.value.error_type == "dealer_mismatch"
        assert any(m.field == "gstin" for m in exc.value.mismatches)

    def test_reject_financial_year_mismatch(self):
        records = [
            ("a.xlsx", DealerMetadata(gstin="03AABCU9603R1ZX", financial_year="2022-23")),
            ("b.xlsx", DealerMetadata(gstin="03AABCU9603R1ZX", financial_year="2023-24")),
        ]
        with pytest.raises(DealerValidationError) as exc:
            validate_dealer_consistency(records)
        assert any(m.field == "financial_year" for m in exc.value.mismatches)

    def test_reject_missing_gstin(self):
        records = [("a.xlsx", DealerMetadata(financial_year="2022-23"))]
        with pytest.raises(DealerValidationError) as exc:
            validate_dealer_consistency(records)
        assert exc.value.error_type == "dealer_metadata_missing"
