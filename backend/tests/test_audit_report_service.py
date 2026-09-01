"""Tests for full audit report service."""

import base64

import openpyxl
import pytest

from models.audit_session import AuditSession
from models.dealer_metadata import DealerMetadata
from services.audit_session_store import clear_sessions, upsert_session
from services.audit_report_service import build_full_docx_report, build_full_excel_report, build_full_pdf_report, build_report_preview
from services.comparison_service import run_gstr1_eway_comparison
from services.comparison_store import clear_session as clear_cmp
from services.investigation_store import clear_session as clear_inv
from tests.comparison_fixtures import build_eway_comparison_workbook, build_gstr1_comparison_workbook

SESSION = "session_report_test"
DEALER = DealerMetadata(gstin="03AABCU9603R1ZX", legal_name="TEST CO", financial_year="2023-24")


@pytest.fixture(autouse=True)
def clean():
    clear_sessions()
    clear_cmp(SESSION)
    clear_inv(SESSION)
    yield
    clear_sessions()


def _session_with_comparison():
    session = AuditSession(session_id=SESSION, dealer=DEALER, financial_year="2023-24")
    upsert_session(session)
    run_gstr1_eway_comparison(
        SESSION,
        gstr1_workbook_base64=base64.b64encode(build_gstr1_comparison_workbook()).decode(),
        ewb_outward_workbook_base64=base64.b64encode(build_eway_comparison_workbook()).decode(),
    )
    return session


class TestAuditReportService:
    def test_report_preview_sections(self):
        session = _session_with_comparison()
        preview = build_report_preview(session)
        assert preview["executive_summary"]["gstin"] == DEALER.gstin
        assert preview["comparison_summary"]["matched_count"] >= 0
        assert len(preview["recommendations"]) >= 1
        assert preview["case_tracking"]["total"] > 0

    def test_excel_report_has_sheets(self):
        session = _session_with_comparison()
        from services.investigation_service import sync_cases_from_comparison
        cases = sync_cases_from_comparison(SESSION)
        buf = build_full_excel_report(session, cases)
        wb = openpyxl.load_workbook(buf)
        assert "Executive Summary" in wb.sheetnames
        assert "Observations" in wb.sheetnames
        assert "Cases" in wb.sheetnames

    def test_pdf_report_valid(self):
        session = _session_with_comparison()
        from services.investigation_service import sync_cases_from_comparison
        cases = sync_cases_from_comparison(SESSION)
        buf = build_full_pdf_report(session, cases)
        assert buf.getvalue().startswith(b"%PDF")

    def test_docx_report_valid(self):
        session = _session_with_comparison()
        from services.investigation_service import sync_cases_from_comparison
        cases = sync_cases_from_comparison(SESSION)
        buf = build_full_docx_report(session, cases)
        assert buf.getvalue()[:2] == b"PK"
